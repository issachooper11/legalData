# 开发者:小白菜
# 开发时间: 2022/3/24 12:11
from openpyxl import load_workbook
import pandas as pd

from util import excel_url, driver_url, originalUrl, excel_url1, driver_url1, originalUrl1
import time
import re


# 获取信息表所有的公司名称和新抓取的数据进行对比
def getCoureCode():
    wb = load_workbook(chooseUrl(3))
    ws = wb['2023信息']
    arr = []
    for num, j in enumerate(ws.rows):
        if num > 0:
            for index, n in enumerate(j):
                if 2 == index:
                    arr.append(n.value)
    return arr


def getCompanyNameAndCode():
    df = pd.read_excel(chooseUrl(3), '2023信息')
    result = []
    for index, row in df.iterrows():
        col2 = row[1]
        col3 = row[2]
        result.append([col2, col3])
    return result


# 根据标记号获取当日呼叫数据并过滤座机号
def getDailyPhoneCallInfo(day_num):
    df = pd.read_excel(chooseUrl(3), '2023信息')
    result = []
    for index, row in df.iterrows():
        if row[0] == float(day_num):
            col2 = row[1]
            col3 = row[2]
            col4 = get_corrcet_phone_number(row[3])
            result.append([col2, col3, col4])
    return result


# 处理上述方法中待呼叫公司电话的数据
def get_corrcet_phone_number(col):
    if isinstance(col, str):
        return re.compile(r'1\d{10}').findall(col)
    else:
        return re.compile(r'1\d{10}').findall(str(col))


# 获取没有验证的所有案号
def getUnEditCodes():
    wb = load_workbook(chooseUrl(3))
    ws = wb['2022信息']
    arr = []
    for num, j in enumerate(ws.rows):
        if num > 0:
            for index, n in enumerate(j):
                if 0 == index:
                    if n.value is not None:
                        break
                if 1 == index:
                    arr.append(n.value)
                if 3 == index:
                    if len(str(n.value)) != 11:
                        arr.pop()
                        break
    return arr


# 获取案号和公司名称
def getCodeAndNames():
    wb = load_workbook(chooseUrl(3))
    ws = wb['2022信息']
    lists = []
    for num, j in enumerate(ws.rows):
        arr = []
        if num > 0:
            for index, n in enumerate(j):
                if 1 == index:
                    arr.append(n.value)
                if 2 == index:
                    arr.append(n.value)
            lists.append(arr)
    return lists


# 对新表循环获取处理过的公司数据
def getFreshData():
    url = chooseUrl(1)
    wb = load_workbook(url)
    ws = wb['Sheet']
    arr = getCoureCode()
    news = wb.create_sheet()
    news.append(['公司名称', '案号', '法院', '日期'])
    for num, j in enumerate(ws.rows):
        if num > 0:
            new_list = []
            for index, n in enumerate(j):
                # 判断顺序 是0则为公司字符串 否则为案号等其他信息
                if 0 == index:
                    if n.value in arr:
                        print('第' + str(num) + '行原表里已经有了...........' + n.value)
                        break
                    else:
                        print('第' + str(num) + '行添到新的数据了...........' + n.value)
                        new_list.append(n.value)
                else:
                    new_list.append(n.value)
            if len(new_list) != 0:
                news.append(new_list)
    wb.save(url)
    print('成功筛选')


# 对抓取完的初始数据进行筛查和处理 找出公司客户
def getNewWkData():
    url = chooseUrl(1)
    wb = load_workbook(url)
    ws = wb['Sheet']
    news = wb.create_sheet()
    news.append(['公司名称', '案号', '法院', '日期'])
    for num, j in enumerate(ws.rows):
        if num > 0:
            list = []
            for index, n in enumerate(j):
                # 判断顺序 是0则为公司字符串 否则为案号等其他信息
                if 0 == index:
                    name = n.value
                    if name.find('、') != -1:
                        name = name.split('、')[0]
                    elif name.find('与') != -1:
                        name = name.split('与')[0]
                        if name.find('等') != -1:
                            name = name.split('等')[0]
                    elif name.find('等') != -1:
                        name = name.split('等')[0]
                    elif name.find('二审') != -1:
                        name = name.split('二审')[0]
                    else:
                        pass
                    if len(name) > 5:
                        if checkInfoValue(name):
                            list.append(name)
                    else:
                        break
                else:
                    list.append(n.value)
            if len(list) == 4:
                news.append(list)
    wb.save(url)
    print('成功筛选')
    a = input('手动剔除重复项目，是否再启动对比程序并添加新数据:')
    # 不符合要求的关闭窗口打开下一个
    if a == '':
        getFreshData()
    else:
        pass


def checkValue(data):
    if '银行' in data:
        return False
    elif '委员会' in data:
        return False
    elif '解放军' in data:
        return False
    elif '律师' in data:
        return False
    elif '学校' in data:
        return False
    elif '医院' in data:
        return False
    elif '政府' in data:
        return False
    elif '学院' in data:
        return False
    elif '大学' in data:
        return False
    elif '中学' in data:
        return False
    elif '小学' in data:
        return False
    elif '合作社' in data:
        return False
    elif '纠纷' in data:
        return False
    elif '×' in data:
        return False
    else:
        return True


def checkInfoValue(data):
    info_list = ['银行', '委员会', '解放军', '律师', '学校', '医院', '政府', '学院', '大学', '中学', '小学', '合作社',
                 '纠纷', '×', '合作联社', '卫生院', '管理段']
    for field in info_list:
        if field in data:
            return False
    return True


# 处理数组内容中含'等 、 二审等字样的内容只截取到原告信息进行判断
# 去重复项目 去空项目 去个人项目
def getPlaintiff(arr):
    new_arr = arr
    separators = ['与', '等', '二审', '、', ';', '*', '·', ',']
    for i in arr:
        for sep in separators:
            i[0] = i[0].split(sep)[0]
    unique_dict = {}
    for array in new_arr:
        key = array[0]
        if key != '':
            if len(key) > 5:
                if checkInfoValue(key):
                    if key not in unique_dict:
                        unique_dict[key] = array
    return list(unique_dict.values())


# 直接将数据转化为excel输出到桌面
def transToExcel(l):
    # 对最终数据再处理一遍
    final_arr = getPlaintiff(l)
    if len(final_arr) > 0:
        save_url = chooseUrl(1)
        wb = load_workbook(save_url)
        news = wb.create_sheet()
        news.append(['公司名称', '案号', '法院', '日期'])
        for i in final_arr:
            news.append(i)
        wb.save(save_url)
        showInfo('转换完成')


# 倒计时
def delay(num):
    while num > 0:
        time.sleep(1)
        showInfo(f'倒计时:{num}s')
        num -= 1


# 打印信息
def showInfo(data):
    print('*' * 11 + data + '*' * 11)


# 人工判断是哪个平台运行的程序,type:1-excel_url 2-driver_url 3-originalUrl
def chooseUrl(typeNum):
    a = input('（选择平台，默认mac路径，否则是win路径）-----------：')
    if a == '':
        if typeNum == 1:
            return excel_url
        elif typeNum == 2:
            return driver_url
        else:
            return originalUrl
    else:
        if typeNum == 1:
            return excel_url1
        elif typeNum == 2:
            return driver_url1
        else:
            return originalUrl1


# 对法院信息重组
def filterCourtName(court):
    if len(court) > 0:
        lists = court.split(' ')
        return [lists[1][1:-1], lists[0][1:-1], lists[2][1:-1]]


# 获取网页抓取的数据
def filterHtmlData(l):
    filter_list = []
    if len(l) > 0:
        for i in l:
            lists = []
            a = i.split('\n')
            # 取数组的第一个 是网页数据内容
            if len(a) > 1:
                b = a[1]
                c = filterCourtName(a[2])
                if True:
                    lists.append(b)
                    for x in c:
                        lists.append(x)
                    filter_list.append(lists)
        print(filter_list)
    if len(filter_list) > 0:
        save_url = chooseUrl(1)
        wb = load_workbook(save_url)
        news = wb.create_sheet()
        news.append(['公司名称', '案号', '法院', '日期'])
        for i in filter_list:
            news.append(i)
        wb.save(save_url)
        showInfo('转换完成')


# 处理微科抓取的数据 整理成判决书内容-案号-法院-日期的数组形式
def changeData(l):
    filter_list = []
    if len(l) > 0:
        for i in l:
            lists = []
            a = i.split('\n')
            # 取数组的第一个 是网页数据内容
            if len(a) > 1:
                b = a[1]
                c = filterCourtName(a[2])
                if True:
                    lists.append(b)
                    for x in c:
                        lists.append(x)
                    filter_list.append(lists)
        return filter_list


# 处理裁判文书网抓取的数据
def change_cpwsw_data(info):
    arr = []
    for i in info:
        lines = i.split('\n')
        data = [lines[1], lines[2].split()[1], lines[2].split()[0], lines[2].split()[2]]
        arr.extend([data])
    return arr


def manual_confirm(data):
    a = input(data)
    if a == '':
        pass
