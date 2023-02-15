import time
from openpyxl import load_workbook

excel_url = "/Users/alex/Desktop/12.xlsx"
excel_url1 = 'C:/Users/money/Desktop/12.xlsx'
login_url = 'http://www.mxstsg.com/account/login'
aiqicha_url = 'https://aiqicha.baidu.com/'
wk_url = 'http://www.mxstsg.com/account/login'
originalUrl = '/Users/alex/Desktop/2023客户数据.xlsx'
originalUrl1 = 'C:/Users/money/Desktop/2023客户数据.xlsx'
driver_url = '/Users/alex/Downloads/edgedriver_mac64/msedgedriver'
driver_url1 = 'E:\software\msedgedriver.exe'

docs = {
    '账号': '//*[@id="login_username"]',
    '密码': '//*[@id="loginform"]/div[3]/div/div/input',
    '登录': '//*[@id="loginform"]/div[6]/div/div[4]/button',
    '法律数据库': '//*[@id="sidebar"]/ul/li[3]/ul/li[3]/a',
    '(6)  威科先行': '//*[@id="content"]/div[2]/div/ul/li[5]/a',
    '入口0': '//*[@id="content"]/div[2]/div/ul/li[1]/a',
    '入口1': '//*[@id="content"]/div[2]/div/ul/li[2]/a',
    '入口2': '//*[@id="content"]/div[2]/div/ul/li[3]/a',
    '入口3': '//*[@id="content"]/div[2]/div/ul/li[4]/a',
    '点击这里登陆——法律信息库': '/html/body/font/b/h3/form/input',
    'entry': '//*[@id="information1"]/table/tbody[1]/tr/td[2]/a/button',
    '案例': '//*[@id="dropdownMenu3"]',
    '裁判文书': '//*[@id="navbar"]/ul/li[3]/ul/li[1]/a',
    '输入框': '//*[@id="queryString"]',
    '点击案件': '//*[@id="column-list"]/li[1]/b-list-item/div/div[1]/b-lock/a/span',
    '搜索': '//*[@id="search_button"]',
    '删除': '/html/body/bold-app/b-judgment/div[1]/div/div/div/div[1]/div[1]/div/div/a',
    '新搜索': '//*[@id="new_common_search_button"]',
    '内容': '//*[@id="main_detail"]',
    '下拉菜单': '/html/body/bold-app/b-judgment/div[2]/b-no-data/div[1]/div/div[2]/div/div[2]/select[2]',
    '列表': '//*[@id="column-list"]',
    '下一页': '/html/body/bold-app/b-judgment/div[2]/b-no-data/div[1]/div/div[2]/div/div[3]/pagination/ul/li[9]',
    '查看更多': '//*[@id="court_more_a"]',
    '二审判决内容': '//*[@id="main_detail"]/span/div/div[3]/div[1]/div[2]',
    '点击进入威科FF入口': '/html/body/div/p[1]/strong/span/a'
}
docsItem = {
    '加号': '//*[@id="causeOfActionǁ01000000000000民事ǁǂ"]/i',
    '合同': '//*[@id="causeOfActionǁ01000000000000民事/01040000000000合同、准合同纠纷ǁǂ_anchor"]',
    '二审': '//*[@id="instanceǁ002ǁǂ_anchor"]',
    '判决书': '//*[@id="typeOfDecisionǁ001ǁǂ_anchor"]',
}

provinces = {
    '浙江省': '//*[@id="courtǁ032000000浙江省ǁǂ_anchor"]',
    '安徽省': '//*[@id="courtǁ002000000安徽省ǁǂ_anchor"]',
    '山东省': '//*[@id="courtǁ023000000山东省ǁǂ_anchor"]',
    '河北省': '//*[@id="courtǁ011000000河北省ǁǂ_anchor"]',
    '江苏省': '//*[@id="courtǁ017000000江苏省ǁǂ_anchor"]',
    '河南省': '//*[@id="courtǁ013000000河南省ǁǂ_anchor"]',
    '北京市': '//*[@id="courtǁ003000000北京市ǁǂ_anchor"]',
    '江西省': '//*[@id="courtǁ018000000江西省ǁǂ_anchor"]',
    '陕西省': '//*[@id="courtǁ026000000陕西省ǁǂ_anchor"]',
    '山西省': '//*[@id="courtǁ025000000山西省ǁǂ_anchor"]',
    '湖北省': '//*[@id="courtǁ014000000湖北省ǁǂ_anchor"]',
    '湖南省': '//*[@id="courtǁ015000000湖南省ǁǂ_anchor"]'
}


# 倒计时
def delay(num):
    while num > 0:
        time.sleep(1)
        showInfo(f'倒计时:{num}s')
        num -= 1


# 打印信息
def showInfo(data):
    print('.' * 11 + data + '.' * 11)


# 人工判断是哪个平台运行的程序,type:1-excel_url 2-driver_url 3-originalUrl
def chooseUrl(typeNum):
    a = input('（选择平台，默认mac路径，否则是win路径）-----------：')
    if a == '':
        if typeNum == 1:
            return excel_url
        elif typeNum == 2:
            return driver_url
        else:
            return originalUrl
    else:
        if typeNum == 1:
            return excel_url1
        elif typeNum == 2:
            return driver_url1
        else:
            return originalUrl1


# 对法院信息重组
def filterCourtName(court):
    if len(court) > 0:
        lists = court.split(' ')
        return [lists[1][1:-1], lists[0][1:-1], lists[2][1:-1]]


# 获取网页抓取的数据
def filterHtmlData(l):
    filter_list = []
    if len(l) > 0:
        for i in l:
            lists = []
            a = i.split('\n')
            # 取数组的第一个 是网页数据内容
            if len(a) > 1:
                b = a[1]
                c = filterCourtName(a[2])
                if True:
                    lists.append(b)
                    for x in c:
                        lists.append(x)
                    filter_list.append(lists)
        print(filter_list)
    if len(filter_list) > 0:
        save_url = chooseUrl(1)
        wb = load_workbook(save_url)
        news = wb.create_sheet()
        news.append(['公司名称', '案号', '法院', '日期'])
        for i in filter_list:
            news.append(i)
        wb.save(save_url)
        showInfo('转换完成')


# 直接将数据转化为excel输出到桌面
def transToExcel(l):
    if len(l) > 0:
        save_url = chooseUrl(1)
        wb = load_workbook(save_url)
        news = wb.create_sheet()
        news.append(['公司名称', '案号', '法院', '日期'])
        for i in l:
            news.append(i)
        wb.save(save_url)
        showInfo('转换完成')


def changeData(l):
    filter_list = []
    if len(l) > 0:
        for i in l:
            lists = []
            a = i.split('\n')
            # 取数组的第一个 是网页数据内容
            if len(a) > 1:
                b = a[1]
                c = filterCourtName(a[2])
                if True:
                    lists.append(b)
                    for x in c:
                        lists.append(x)
                    filter_list.append(lists)
        return filter_list


# 手动点击通用方法
def handleByClick():
    a = input('是否继续')
    if a == '':
        pass


def handleByInfo(data):
    a = input('...........等待页面加载...........')
    if a == '':
        print('.' * 11 + data + '.' * 11)
        pass


def manual_confirm(data):
    a = input(data)
    if a == '':
        pass
